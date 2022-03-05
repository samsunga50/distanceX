from openpyxl import load_workbook
from distance import distance

travel = load_workbook('travel.xlsx')
ws = travel.active

row = 2
#while cell has a value
while(ws["A" + str(row)].value != None):
    origin = ws["A" + str(row)].value
    destination = ws["B" + str(row)].value

    km = distance(origin, destination)
    print(origin + " to " + destination + ": " + km + " km")

    ws["C" + str(row)].value = km
    row = row + 1
    
travel.save('travel.xlsx')